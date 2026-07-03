"""
cuesheet.py — Pro Tools text export → CSV + branded PDF.

CLI:
    python3 cuesheet.py <path-to-protools-export.txt>

Importable:
    import cuesheet
    csv_path, pdf_path = cuesheet.main("/path/to/session.txt")
"""
import os
import re
import sys
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer, Image, Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


HEADER_LABELS = [
    'SESSION NAME:',
    'SAMPLE RATE:',
    'BIT DEPTH:',
    'SESSION START TIMECODE:',
    'TIMECODE FORMAT:',
]

MERGE_GAP_SECONDS = 5   # default gap for consolidating cues

# Timecode format for the cue sheet, set from the session's TIMECODE FORMAT
# header when the input is read. Defaults to 25fps non-drop.
_FPS = 25
_DROP = False


def set_tc_format(fps_int, drop):
    global _FPS, _DROP
    _FPS, _DROP = fps_int, drop


def parse_tc_format(text):
    """'29.97 Drop Frame' -> (30, True); '25 Frame' -> (25, False)."""
    low = text.lower()
    drop = 'drop' in low and 'non' not in low
    m = re.search(r'[\d.]+', text)
    return (round(float(m.group())) if m else 25), drop


def clean_name(name):
    return re.sub(r'\.\d{2}(-\d{2})?$', '', name)


def frames_to_tc(frames):
    """Sequential frame count -> HH:MM:SS:FF (';FF' when drop-frame)."""
    fps = _FPS
    if _DROP:
        dropf = 2 * (fps // 30)
        fp10 = fps * 600 - dropf * 9
        fpm = fps * 60 - dropf
        n = frames % (fps * 3600 * 24)
        d, m = divmod(n, fp10)
        n += dropf * 9 * d + (dropf * ((m - dropf) // fpm) if m > dropf else 0)
        sep = ';'
    else:
        n, sep = frames, ':'
    return (f"{(n // (fps * 3600)) % 24:02d}:{(n // (fps * 60)) % 60:02d}:"
            f"{(n // fps) % 60:02d}{sep}{n % fps:02d}")


def tc_to_frames(tc):
    """HH:MM:SS:FF (or ';FF') -> sequential frame count."""
    h, m, s, f = (int(x) for x in re.split(r'[:;]', tc.strip()))
    fps = _FPS
    total = (h * 3600 + m * 60 + s) * fps + f
    if _DROP:
        mins = h * 60 + m
        total -= 2 * (fps // 30) * (mins - mins // 10)
    return total


def common_prefix(a, b):
    i = 0
    while i < len(a) and i < len(b) and a[i] == b[i]:
        i += 1
    return a[:i]


def overlaps_or_within(start_a, end_a, start_b, end_b, gap_frames):
    a_start = tc_to_frames(start_a)
    a_end   = tc_to_frames(end_a)
    b_start = tc_to_frames(start_b)
    b_end   = tc_to_frames(end_b)
    return a_start - gap_frames < b_end and b_start - gap_frames < a_end


def cleanup_display_name(prefix):
    prefix = prefix.rstrip()
    m = re.search(r'[ _]\d+$', prefix)
    if m:
        prefix = prefix[:m.start()]
    return prefix.rstrip(' _')


def read_file(input_file):
    """Parse a Pro Tools text export, returning (clips, session_info)."""
    clips = []
    session_info = {}
    with open(input_file, 'r', encoding='UTF-8') as f:
        for line in f:
            if not line.strip():
                continue
            parts = line.split('\t')
            if len(parts) == 2 and parts[0] in HEADER_LABELS:
                session_info[parts[0]] = parts[1].strip()
                if parts[0] == 'TIMECODE FORMAT:':
                    set_tc_format(*parse_tc_format(parts[1]))
                continue
            if len(parts) < 7:
                continue
            name     = parts[2].strip()   # raw name; cleaned later for the cue sheet
            start    = parts[3].strip()
            end      = parts[4].strip()
            duration = parts[5].strip()
            state    = parts[6].strip()
            if not re.match(r'^\d{2}:\d{2}:\d{2}[:;]\d{2}$', start):
                continue
            if not re.match(r'^\d{2}:\d{2}:\d{2}[:;]\d{2}$', end):
                continue
            if state != 'Unmuted':
                continue
            if start == end:
                continue
            clips.append((name, start, end, duration))
    return clips, session_info


def write_pdf(clips, output_path, session_info, logo_path):
    BRAND_DARK = colors.HexColor('#1a1a1a')
    LABEL_GREY = colors.HexColor('#666666')
    DIVIDER    = colors.HexColor('#e0e0e0')
    ROW_ALT    = colors.HexColor('#fafafa')

    doc = SimpleDocTemplate(output_path, pagesize=A4,
        leftMargin=15*mm, rightMargin=15*mm, topMargin=15*mm, bottomMargin=18*mm,
        title="Music Cue Sheet")
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=20, textColor=BRAND_DARK, leading=22, spaceAfter=2)
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontName='Helvetica', fontSize=11, textColor=LABEL_GREY, leading=14)
    label_style = ParagraphStyle('Label', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=7, textColor=LABEL_GREY, leading=10)
    value_style = ParagraphStyle('Value', parent=styles['Normal'], fontName='Helvetica', fontSize=8.5, textColor=BRAND_DARK, leading=10)
    cell_name_style = ParagraphStyle('CellName', parent=styles['Normal'], fontName='Helvetica', fontSize=7.5, textColor=BRAND_DARK, leading=9.5)

    title_block = Table([[Paragraph('MUSIC CUE SHEET', title_style)], [Paragraph(session_info.get('SESSION NAME:', ''), subtitle_style)]], colWidths=[145*mm])
    title_block.setStyle(TableStyle([('LEFTPADDING', (0,0), (-1,-1), 0), ('RIGHTPADDING', (0,0), (-1,-1), 0), ('TOPPADDING', (0,0), (-1,-1), 0), ('BOTTOMPADDING', (0,0), (-1,-1), 0)]))

    if os.path.exists(logo_path):
        logo = Image(logo_path, width=22*mm, height=22*mm)
        header_strip = Table([[title_block, logo]], colWidths=[145*mm, 35*mm])
    else:
        header_strip = Table([[title_block]], colWidths=[180*mm])
    header_strip.setStyle(TableStyle([('VALIGN', (0,0), (-1,0), 'TOP'), ('ALIGN', (1,0), (1,0), 'RIGHT'), ('LEFTPADDING', (0,0), (-1,-1), 0), ('RIGHTPADDING', (0,0), (-1,-1), 0), ('TOPPADDING', (0,0), (-1,-1), 0), ('BOTTOMPADDING', (0,0), (-1,-1), 0)]))

    label_map = {'SESSION NAME:': 'SESSION NAME', 'SAMPLE RATE:': 'SAMPLE RATE', 'BIT DEPTH:': 'BIT DEPTH', 'SESSION START TIMECODE:': 'START TC', 'TIMECODE FORMAT:': 'FRAME RATE'}
    info_row = []
    for label in HEADER_LABELS:
        info_row.append([Paragraph(label_map[label], label_style), Paragraph(session_info.get(label, ''), value_style)])
    info_table = Table([info_row], colWidths=[36*mm]*5)
    info_table.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'TOP'), ('LEFTPADDING', (0,0), (-1,-1), 0), ('RIGHTPADDING', (0,0), (-1,-1), 6), ('TOPPADDING', (0,0), (-1,-1), 8), ('BOTTOMPADDING', (0,0), (-1,-1), 8), ('LINEABOVE', (0,0), (-1,0), 0.5, DIVIDER), ('LINEBELOW', (0,0), (-1,0), 0.5, DIVIDER)]))

    table_data = [['No.', 'Name', 'Timecode In', 'Timecode Out', 'Duration']]
    for i, clip in enumerate(clips, start=1):
        name, start, end, duration = clip
        table_data.append([str(i), Paragraph(name, cell_name_style), start, end, duration])
    col_widths = [11*mm, 95*mm, 24*mm, 24*mm, 26*mm]
    cue_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    ts = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), BRAND_DARK), ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'), ('FONTSIZE', (0,0), (-1,0), 8),
        ('ALIGN', (0,0), (1,0), 'LEFT'), ('ALIGN', (2,0), (-1,0), 'RIGHT'),
        ('TOPPADDING', (0,0), (-1,0), 7), ('BOTTOMPADDING', (0,0), (-1,0), 7),
        ('LEFTPADDING', (0,0), (-1,0), 6), ('RIGHTPADDING', (0,0), (-1,0), 6),
        ('FONTNAME', (0,1), (-1,-1), 'Helvetica'), ('FONTSIZE', (0,1), (-1,-1), 7.5),
        ('TEXTCOLOR', (0,1), (-1,-1), BRAND_DARK), ('VALIGN', (0,1), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,1), (-1,-1), 4), ('BOTTOMPADDING', (0,1), (-1,-1), 4),
        ('LEFTPADDING', (0,1), (-1,-1), 6), ('RIGHTPADDING', (0,1), (-1,-1), 6),
        ('ALIGN', (0,1), (0,-1), 'RIGHT'), ('ALIGN', (1,1), (1,-1), 'LEFT'),
        ('ALIGN', (2,1), (-1,-1), 'RIGHT'), ('FONTNAME', (2,1), (-1,-1), 'Courier'),
        ('LINEBELOW', (0,0), (-1,-1), 0.25, DIVIDER),
    ])
    for row in range(1, len(table_data)):
        if row % 2 == 0:
            ts.add('BACKGROUND', (0, row), (-1, row), ROW_ALT)
    cue_table.setStyle(ts)

    def draw_footer(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 7.5)
        canvas.setFillColor(LABEL_GREY)
        canvas.drawString(15*mm, 10*mm, 'Mighty Sound  ·  mightysound.studio')
        canvas.drawRightString(A4[0] - 15*mm, 10*mm, f'Page {canvas.getPageNumber()}')
        canvas.restoreState()

    doc.build([header_strip, Spacer(1, 6*mm), info_table, Spacer(1, 8*mm), cue_table], onFirstPage=draw_footer, onLaterPages=draw_footer)


XLSX_LABEL_MAP = {
    'SESSION NAME:': 'Session Name',
    'SAMPLE RATE:': 'Sample Rate',
    'BIT DEPTH:': 'Bit Depth',
    'SESSION START TIMECODE:': 'Start TC',
    'TIMECODE FORMAT:': 'Frame Rate',
}
BRAND_DARK_HEX = '1A1A1A'
LABEL_GREY_HEX = '666666'


def _fill_sheet(ws, session_info, clips):
    """Lay out one worksheet: title + session-info block + cue table."""
    ws['A1'] = 'MUSIC CUE SHEET'
    ws['A1'].font = Font(bold=True, size=16, color=BRAND_DARK_HEX)
    ws['A2'] = session_info.get('SESSION NAME:', '')
    ws['A2'].font = Font(size=11, color=LABEL_GREY_HEX)

    r = 4
    for label in HEADER_LABELS:
        if label == 'TIMECODE FORMAT:':
            continue  # frame rate omitted — confusing on delivery
        lab = ws.cell(row=r, column=1, value=XLSX_LABEL_MAP.get(label, label))
        lab.font = Font(bold=True, size=9, color=LABEL_GREY_HEX)
        val = ws.cell(row=r, column=2, value=session_info.get(label, ''))
        val.font = Font(size=10, color=BRAND_DARK_HEX)
        r += 1

    header_row = r + 1
    headers = ['No.', 'Name', 'Timecode In', 'Timecode Out', 'Duration']
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF', size=10)
        cell.fill = PatternFill('solid', fgColor=BRAND_DARK_HEX)
        cell.alignment = Alignment(horizontal=('left' if col <= 2 else 'right'),
                                   vertical='center')

    row = header_row + 1
    for i, clip in enumerate(clips, start=1):
        name, start, end, duration = clip
        ws.cell(row=row, column=1, value=i).alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=2, value=name)
        for col, val in ((3, start), (4, end), (5, duration)):
            c = ws.cell(row=row, column=col, value=val)
            c.number_format = '@'  # keep timecodes as text, not auto-parsed dates
            c.alignment = Alignment(horizontal='right')
            c.font = Font(name='Courier New', size=10)
        row += 1

    for col, width in {1: 6, 2: 62, 3: 15, 4: 15, 5: 13}.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # Freeze the title/info/header rows so the table header stays on screen.
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def write_xlsx(merged_clips, raw_clips, output_path, session_info):
    """Write one workbook with two sheets: 'Cue Sheet' (merged/grouped) and
    'Raw Clips' (every parsed clip, completely unedited)."""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = 'Cue Sheet'
    _fill_sheet(ws1, session_info, merged_clips)
    ws2 = wb.create_sheet('Raw Clips')
    _fill_sheet(ws2, session_info, raw_clips)
    wb.save(output_path)


_LIB_KEY = re.compile(r'^([A-Za-z]{2,}\d+_\d+)')


def cue_key(name):
    """Identity of the underlying cue. Library music is named
    ALBUM_TRACK_VARIANT TITLE_STEM (e.g. 'XRC139_07_9 ALL IN THE DAYS WORK_STRINGS'),
    so every variant and stem of one track shares its ALBUM_TRACK code
    ('XRC139_07'). Anything without that catalogue code keys on its own name —
    we do NOT fuzzy-match differently-named clips.
    """
    m = _LIB_KEY.match(name)
    return m.group(1) if m else name


def _title_of(name, key):
    """Human title of a library clip: text after the ALBUM_TRACK_VARIANT code,
    minus the file-extension / version tail."""
    rest = name[len(key):]
    rest = re.sub(r'^_\d+', '', rest).lstrip('_ ')                 # drop _VARIANT
    rest = re.sub(r'\.(wav|mp3|aif|aiff|new)\b.*$', '', rest, flags=re.I)
    return rest.strip()


def cue_display_name(members, key):
    """Name for a consolidated cue. Library cues use the common title across
    their members (so stems/variants collapse to the track title); everything
    else uses its own cleaned name."""
    if not _LIB_KEY.fullmatch(key):
        return clean_name(members[0])
    titles = [_title_of(m, key) for m in members] or ['']
    common = titles[0]
    for t in titles[1:]:
        n = 0
        while n < len(common) and n < len(t) and common[n] == t[n]:
            n += 1
        common = common[:n]
    common = re.sub(r'[ _]+$', '', common).strip()
    return common or titles[0] or clean_name(members[0])


def main(input_file, merge_gap_seconds=None):
    """Run the full pipeline. Returns (xlsx_path, pdf_path).

    Writes a single .xlsx workbook (sheet 'Cue Sheet' matches the PDF; sheet
    'All Clips' is the raw deduped list) plus the branded PDF.

    merge_gap_seconds: gap (in seconds) used to consolidate cues. Defaults to
    MERGE_GAP_SECONDS (5s). Converted to frames using the session's actual
    frame rate (read from the TIMECODE FORMAT header).
    """
    input_dir = os.path.dirname(input_file)
    input_name = os.path.basename(input_file)
    base_name = os.path.splitext(input_name)[0]
    xlsx_output = os.path.join(input_dir, f"{base_name}_cuesheet.xlsx")

    # read_file sets the timecode format (_FPS/_DROP) from the header first,
    # so the gap and all TC math below use the session's real frame rate.
    clips, session_info = read_file(input_file)
    gap = MERGE_GAP_SECONDS if merge_gap_seconds is None else merge_gap_seconds
    gap_frames = round(gap * _FPS)
    print(f"Parsed {len(clips)} clips")

    # Normalise names (strip version suffixes like .01) so stereo/copy pairs
    # sitting at the same position collapse to one, then sort by start.
    clips = [(clean_name(name), start, end, dur)
             for (name, start, end, dur) in clips]
    clips.sort(key=lambda c: c[1])

    # Dedupe identical (name, start, end) — collapses stereo L/R pairs
    seen = set()
    unique_clips = []
    for clip in clips:
        name, start, end, duration = clip
        key = (name, start, end)
        if key in seen:
            continue
        seen.add(key)
        unique_clips.append(clip)
    clips = unique_clips

    # Raw Clips sheet = every DISTINCT clip (deduped, so stereo/copy pairs at
    # the same position don't show twice), unmerged and ungrouped.
    raw_clips = list(unique_clips)

    # Consolidate clips into cues. A cue = one library track (ALBUM_TRACK, so
    # all its variants/stems fold together) or, for anything without a
    # catalogue code, one exact name — NO fuzzy matching. Same-cue clips that
    # overlap or sit within the gap merge into one entry even when other clips
    # are interleaved between them in time (the open cue is tracked per key, so
    # a music cue chopped into segments stays one row).
    open_cues = {}
    cues = []   # each: [key, start, end, [member names]]
    for name, start, end, duration in clips:  # sorted by start
        key = cue_key(name)
        idx = open_cues.get(key)
        if idx is not None and tc_to_frames(start) - tc_to_frames(cues[idx][2]) <= gap_frames:
            if tc_to_frames(end) > tc_to_frames(cues[idx][2]):
                cues[idx][2] = end
            cues[idx][3].append(name)
            continue
        cues.append([key, start, end, [name]])
        open_cues[key] = len(cues) - 1

    clips = [
        (cue_display_name(members, key), start, end,
         frames_to_tc(tc_to_frames(end) - tc_to_frames(start)))
        for key, start, end, members in cues
    ]

    # One workbook, two sheets: 'Cue Sheet' (merged/grouped) and 'Raw Clips'
    # (every clip, completely unedited). A real .xlsx opens cleanly in Excel
    # and Google Sheets.
    write_xlsx(clips, raw_clips, xlsx_output, session_info)
    print(f"Wrote {xlsx_output} ({len(clips)} cues / {len(raw_clips)} raw clips)")

    return xlsx_output


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python3 cuesheet.py <path-to-protools-export.txt>")
        sys.exit(1)
    main(sys.argv[1])
